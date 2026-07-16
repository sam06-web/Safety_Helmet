"""
Report generation service: PDF (ReportLab) and Excel (openpyxl) exports.
"""
import io
from datetime import datetime, timedelta

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session

from app.models import Incident, Worker, Helmet, SensorReading, WorkerStatus


# ═══════════════════════════════════════════════════════════════════════════
# PDF Reports
# ═══════════════════════════════════════════════════════════════════════════

def _build_header(styles):
    """Return report header elements."""
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"],
        fontSize=18, spaceAfter=6, textColor=colors.HexColor("#1a237e"),
    )
    subtitle_style = ParagraphStyle(
        "ReportSubtitle", parent=styles["Normal"],
        fontSize=10, textColor=colors.grey,
    )
    return title_style, subtitle_style


def generate_daily_pdf(db: Session, report_date: datetime) -> io.BytesIO:
    """Generate a daily safety summary PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style, subtitle_style = _build_header(styles)
    elements = []

    day_start = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
    day_end = day_start + timedelta(days=1)

    # Title
    elements.append(Paragraph("Industrial Safety Helmet Monitoring System", title_style))
    elements.append(Paragraph(
        f"Daily Safety Report — {report_date.strftime('%B %d, %Y')}", subtitle_style
    ))
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a237e")))
    elements.append(Spacer(1, 12))

    # Summary counts
    total_workers = db.query(Worker).filter(Worker.is_active == True).count()
    incidents_today = (
        db.query(Incident)
        .filter(Incident.timestamp >= day_start, Incident.timestamp < day_end)
        .all()
    )
    active_workers = db.query(Worker).filter(
        Worker.is_active == True, Worker.assigned_helmet_id.isnot(None)
    ).count()

    # PPE compliance
    ppe_ok = 0
    total_checked = 0
    workers_with_helmets = (
        db.query(Worker)
        .filter(Worker.is_active == True, Worker.assigned_helmet_id.isnot(None))
        .all()
    )
    for w in workers_with_helmets:
        latest = (
            db.query(SensorReading)
            .filter(SensorReading.worker_id == w.id, SensorReading.timestamp < day_end)
            .order_by(SensorReading.timestamp.desc())
            .first()
        )
        if latest:
            total_checked += 1
            if latest.ppe_status:
                ppe_ok += 1
    compliance = round((ppe_ok / total_checked * 100) if total_checked else 100, 1)

    summary_data = [
        ["Metric", "Value"],
        ["Total Workers", str(total_workers)],
        ["Workers with Helmets", str(active_workers)],
        ["Total Incidents", str(len(incidents_today))],
        ["PPE Compliance", f"{compliance}%"],
        ["Report Date", report_date.strftime("%Y-%m-%d")],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Incidents table
    elements.append(Paragraph("Incidents", styles["Heading2"]))
    if incidents_today:
        inc_header = ["Time", "Worker", "Type", "Severity", "Location", "Status"]
        inc_rows = [inc_header]
        for inc in incidents_today[:30]:
            worker = db.query(Worker).filter(Worker.id == inc.worker_id).first()
            inc_rows.append([
                inc.timestamp.strftime("%H:%M:%S"),
                worker.name if worker else "Unknown",
                inc.incident_type.value.replace("_", " ").title(),
                inc.severity.value.upper(),
                inc.location or "—",
                "Resolved" if inc.resolved else ("Ack'd" if inc.acknowledged else "Open"),
            ])
        inc_table = Table(inc_rows, colWidths=[1 * inch, 1.2 * inch, 1.2 * inch, 0.8 * inch, 1.2 * inch, 0.8 * inch])
        inc_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e53935")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(inc_table)
    else:
        elements.append(Paragraph("No incidents reported.", styles["Normal"]))

    elements.append(Spacer(1, 20))

    # Worker attendance
    elements.append(Paragraph("Worker Status Summary", styles["Heading2"]))
    all_workers = db.query(Worker).filter(Worker.is_active == True).all()
    status_rows = [["Employee ID", "Name", "Department", "Status"]]
    for w in all_workers:
        status_rows.append([
            w.employee_id, w.name, w.department, w.status.value.title()
        ])
    att_table = Table(status_rows, colWidths=[1.2 * inch, 1.5 * inch, 1.5 * inch, 1 * inch])
    att_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(att_table)

    # Footer
    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')} — "
        "Safety Helmet Monitoring System v1.0",
        subtitle_style,
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_weekly_pdf(db: Session) -> io.BytesIO:
    """Generate a weekly incident report PDF."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
    styles = getSampleStyleSheet()
    title_style, subtitle_style = _build_header(styles)
    elements = []

    now = datetime.utcnow()
    week_start = now - timedelta(days=7)

    elements.append(Paragraph("Industrial Safety Helmet Monitoring System", title_style))
    elements.append(Paragraph(
        f"Weekly Incident Report — {week_start.strftime('%b %d')} to {now.strftime('%b %d, %Y')}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1a237e")))
    elements.append(Spacer(1, 12))

    incidents = (
        db.query(Incident)
        .filter(Incident.timestamp >= week_start)
        .order_by(Incident.timestamp.desc())
        .all()
    )

    # Summary
    total = len(incidents)
    resolved_count = sum(1 for i in incidents if i.resolved)
    critical_count = sum(1 for i in incidents if i.severity.value == "critical")

    summary_data = [
        ["Metric", "Value"],
        ["Total Incidents", str(total)],
        ["Resolved", str(resolved_count)],
        ["Unresolved", str(total - resolved_count)],
        ["Critical Incidents", str(critical_count)],
    ]
    st = Table(summary_data, colWidths=[3 * inch, 3 * inch])
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 20))

    # Incident breakdown by type
    elements.append(Paragraph("Incidents by Type", styles["Heading2"]))
    from collections import Counter
    type_counts = Counter(i.incident_type.value for i in incidents)
    type_rows = [["Incident Type", "Count"]]
    for t, c in type_counts.most_common():
        type_rows.append([t.replace("_", " ").title(), str(c)])
    tt = Table(type_rows, colWidths=[3 * inch, 3 * inch])
    tt.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ff6f00")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(tt)
    elements.append(Spacer(1, 20))

    # Full incident list (limited)
    elements.append(Paragraph("Incident Details", styles["Heading2"]))
    inc_header = ["Date/Time", "Worker", "Type", "Severity", "Location", "Resolved"]
    inc_rows = [inc_header]
    for inc in incidents[:50]:
        worker = db.query(Worker).filter(Worker.id == inc.worker_id).first()
        inc_rows.append([
            inc.timestamp.strftime("%m/%d %H:%M"),
            worker.name if worker else "Unknown",
            inc.incident_type.value.replace("_", " ").title(),
            inc.severity.value.upper(),
            inc.location or "—",
            "Yes" if inc.resolved else "No",
        ])
    it = Table(inc_rows, colWidths=[1.1 * inch, 1.1 * inch, 1.2 * inch, 0.8 * inch, 1.2 * inch, 0.7 * inch])
    it.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e53935")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(it)

    elements.append(Spacer(1, 30))
    elements.append(Paragraph(
        f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')} — "
        "Safety Helmet Monitoring System v1.0",
        subtitle_style,
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ═══════════════════════════════════════════════════════════════════════════
# Excel Export
# ═══════════════════════════════════════════════════════════════════════════

def _style_header_row(ws, num_cols: int):
    """Apply header styling to the first row."""
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def generate_excel_export(db: Session, data_type: str = "incidents") -> io.BytesIO:
    """Generate an Excel workbook with requested data sheets."""
    wb = Workbook()

    # ── Incidents sheet ──────────────────────────────────────────────────
    ws_inc = wb.active
    ws_inc.title = "Incidents"
    inc_headers = [
        "ID", "Timestamp", "Worker ID", "Worker Name", "Helmet ID",
        "Incident Type", "Severity", "Description", "Gas Level",
        "Temperature", "Location", "Acknowledged", "Resolved",
    ]
    ws_inc.append(inc_headers)
    _style_header_row(ws_inc, len(inc_headers))

    incidents = db.query(Incident).order_by(Incident.timestamp.desc()).limit(1000).all()
    for inc in incidents:
        worker = db.query(Worker).filter(Worker.id == inc.worker_id).first()
        helmet = db.query(Helmet).filter(Helmet.id == inc.helmet_id).first()
        ws_inc.append([
            inc.id,
            inc.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            worker.employee_id if worker else "",
            worker.name if worker else "",
            helmet.helmet_id if helmet else "",
            inc.incident_type.value,
            inc.severity.value,
            inc.description,
            inc.gas_level,
            inc.temperature,
            inc.location or "",
            "Yes" if inc.acknowledged else "No",
            "Yes" if inc.resolved else "No",
        ])

    # Auto-size columns
    for col in ws_inc.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_inc.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # ── Workers sheet ────────────────────────────────────────────────────
    ws_wrk = wb.create_sheet(title="Workers")
    wrk_headers = [
        "ID", "Employee ID", "Name", "Department", "Shift",
        "Status", "Contact Phone", "Helmet Assigned", "Active",
    ]
    ws_wrk.append(wrk_headers)
    _style_header_row(ws_wrk, len(wrk_headers))

    workers = db.query(Worker).order_by(Worker.name).all()
    for w in workers:
        helmet = db.query(Helmet).filter(Helmet.id == w.assigned_helmet_id).first() if w.assigned_helmet_id else None
        ws_wrk.append([
            w.id, w.employee_id, w.name, w.department, w.shift.value,
            w.status.value, w.contact_phone or "",
            helmet.helmet_id if helmet else "None",
            "Yes" if w.is_active else "No",
        ])

    for col in ws_wrk.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_wrk.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    # ── Sensor Data sheet (last 24 hours) ────────────────────────────────
    ws_sensor = wb.create_sheet(title="Sensor Data (24h)")
    sensor_headers = [
        "Timestamp", "Worker", "Helmet", "Temperature (°C)", "Gas (ppm)",
        "PPE", "Buckle", "Emergency", "Battery (%)", "Signal (%)", "Location",
    ]
    ws_sensor.append(sensor_headers)
    _style_header_row(ws_sensor, len(sensor_headers))

    since = datetime.utcnow() - timedelta(hours=24)
    readings = (
        db.query(SensorReading)
        .filter(SensorReading.timestamp >= since)
        .order_by(SensorReading.timestamp.desc())
        .limit(5000)
        .all()
    )
    for r in readings:
        worker = db.query(Worker).filter(Worker.id == r.worker_id).first()
        helmet = db.query(Helmet).filter(Helmet.id == r.helmet_id).first()
        ws_sensor.append([
            r.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
            worker.name if worker else "",
            helmet.helmet_id if helmet else "",
            round(r.temperature, 1),
            round(r.gas_level, 1),
            "Yes" if r.ppe_status else "No",
            "Yes" if r.buckle_status else "No",
            "Yes" if r.emergency_button else "No",
            r.battery_pct,
            r.signal_strength,
            r.location or "",
        ])

    for col in ws_sensor.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws_sensor.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
